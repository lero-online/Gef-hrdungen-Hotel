import json
from dataclasses import dataclass, asdict, field
from datetime import date, datetime
from typing import List, Optional, Dict, Any, Tuple
from dateutil.relativedelta import relativedelta
from io import BytesIO
import re

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule

# =========================
# Datenmodelle
# =========================

STOP_LEVELS = [
    "S (Substitution/Quelle entfernen)",
    "T (Technisch)",
    "O (Organisatorisch)",
    "P (PSA)",
    "Q (Qualifikation/Unterweisung)"
]
STATUS_LIST = ["offen", "in Umsetzung", "wirksam", "nicht wirksam", "entfallen"]

@dataclass
class Measure:
    title: str
    stop_level: str
    responsible: str = ""
    due_date: Optional[str] = None  # ISO
    status: str = "offen"
    notes: str = ""

@dataclass
class Hazard:
    id: str
    area: str
    activity: str
    hazard: str
    sources: List[str]
    existing_controls: List[str]
    prob: int = 3
    sev: int = 3
    risk_value: int = 9
    risk_level: str = "mittel"
    additional_measures: List[Measure] = field(default_factory=list)
    last_review: Optional[str] = None
    reviewer: str = ""
    documentation_note: str = ""

@dataclass
class Assessment:
    company: str
    location: str
    created_at: str
    created_by: str
    industry: str = "Hotel/Gastgewerbe"
    scope_note: str = ""
    risk_matrix_thresholds: Dict[str, List[int]] = field(default_factory=lambda: {"thresholds": [6, 12, 16]})
    hazards: List[Hazard] = field(default_factory=list)
    measures_plan_note: str = ""
    documentation_note: str = ""
    next_review_hint: str = ""

# =========================
# Utility
# =========================

def compute_risk(prob: int, sev: int, thresholds: List[int]) -> Tuple[int, str]:
    v = prob * sev
    if v <= thresholds[0]:
        return v, "niedrig"
    elif v <= thresholds[1]:
        return v, "mittel"
    elif v <= thresholds[2]:
        return v, "hoch"
    else:
        return v, "sehr hoch"

def hazard_to_row(h: Hazard) -> Dict[str, Any]:
    return {
        "ID": h.id, "Bereich": h.area, "Tätigkeit": h.activity, "Gefährdung": h.hazard,
        "Quellen/Einwirkungen": "; ".join(h.sources), "Bestehende Maßnahmen": "; ".join(h.existing_controls),
        "Eintrittswahrscheinlichkeit (1-5)": h.prob, "Schadensschwere (1-5)": h.sev,
        "Risikosumme": h.risk_value, "Risikostufe": h.risk_level,
        "Letzte Prüfung": h.last_review or "", "Prüfer/in": h.reviewer,
        "Dokumentationshinweis": h.documentation_note
    }

def measures_to_rows(h: Hazard) -> List[Dict[str, Any]]:
    rows = []
    for m in h.additional_measures:
        rows.append({
            "Gefährdungs-ID": h.id, "Bereich": h.area, "Gefährdung": h.hazard,
            "Maßnahme": m.title, "STOP(+Q)": m.stop_level, "Verantwortlich": m.responsible,
            "Fällig am": m.due_date or "", "Status": m.status, "Hinweis": m.notes
        })
    return rows

def new_id(prefix="HZ", n=4) -> str:
    ts = datetime.now().strftime("%y%m%d%H%M%S%f")[-n:]
    return f"{prefix}-{int(datetime.now().timestamp())}-{ts}"

def dump_excel(assess: Assessment) -> bytes:
    # --- Datenaufbereitung ---
    hazards_df = pd.DataFrame([hazard_to_row(h) for h in assess.hazards])
    measures_df = pd.DataFrame([r for h in assess.hazards for r in measures_to_rows(h)])

    # Maßnahmen-Plan (Schritt 5) – inkl. Status/Verantwortlich/Fällig
    plan_rows = []
    for h in assess.hazards:
        for m in h.additional_measures:
            plan_rows.append({
                "Gefährdungs-ID": h.id,
                "Bereich": h.area,
                "Tätigkeit": h.activity,
                "Gefährdung": h.hazard,
                "Risikosumme": h.risk_value,
                "Risikostufe": h.risk_level,
                "Maßnahme": m.title,
                "STOP(+Q)": m.stop_level,
                "Verantwortlich": m.responsible,
                "Fällig am": m.due_date or "",
                "Status": m.status,
                "Hinweis": m.notes,
            })
    plan_df = pd.DataFrame(plan_rows)

    # Wirksamkeit (Schritt 6) je Gefährdung
    review_rows = []
    for h in assess.hazards:
        review_rows.append({
            "Gefährdungs-ID": h.id,
            "Bereich": h.area,
            "Tätigkeit": h.activity,
            "Gefährdung": h.hazard,
            "Letzte Prüfung": h.last_review or "",
            "Prüfer/in": h.reviewer,
            "Beurteilungs-/Dokumentationshinweis": h.documentation_note,
        })
    review_df = pd.DataFrame(review_rows)

    # Meta / Stammdaten (Schritt 1)
    meta = {
        "Unternehmen": assess.company,
        "Standort": assess.location,
        "Erstellt am": assess.created_at,
        "Erstellt von": assess.created_by,
        "Branche": assess.industry,
        "Umfang/Scope": assess.scope_note,
    }
    meta_df = pd.DataFrame(list(meta.items()), columns=["Feld", "Wert"])

    # Dokumentation (Schritt 7)
    doc_df = pd.DataFrame({"Dokumentationshinweis": [assess.documentation_note or ""]})

    # Fortschreiben (Schritt 8)
    prog_df = pd.DataFrame({"Anlässe/Fristen (Fortschreibung)": [assess.next_review_hint or ""]})

    # Konfiguration
    thresholds = assess.risk_matrix_thresholds.get("thresholds", [6, 12, 16])
    conf_df = pd.DataFrame(
        {"Einstellung": ["Grenze niedrig (≤)", "Grenze mittel (≤)", "Grenze hoch (≤)"],
         "Wert": [thresholds[0], thresholds[1], thresholds[2]]}
    )

    # --- Excel schreiben ---
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Reihenfolge/Blätter:
        meta_df.to_excel(writer, sheet_name="01_Stammdaten", index=False)
        hazards_df.to_excel(writer, sheet_name="10_Gefährdungen", index=False)
        measures_df.to_excel(writer, sheet_name="20_Maßnahmen", index=False)
        plan_df.to_excel(writer, sheet_name="30_Plan", index=False)
        review_df.to_excel(writer, sheet_name="40_Wirksamkeit", index=False)
        doc_df.to_excel(writer, sheet_name="50_Dokumentation", index=False)
        prog_df.to_excel(writer, sheet_name="60_Fortschreiben", index=False)
        conf_df.to_excel(writer, sheet_name="90_Konfiguration", index=False)

        # README
        readme_text = [
            ["Datei erstellt", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Generator", "Gefährdungsbeurteilung Streamlit-App"],
            ["Hinweis", "Blätter 10–60 bilden die Prozessschritte ab. "
                        "Risikofarben auf Blatt 10 beziehen sich auf die Risikosumme."],
            ["Kontakt", assess.created_by or ""],
        ]
        readme_df = pd.DataFrame(readme_text, columns=["Info", "Wert"])
        readme_df.to_excel(writer, sheet_name="99_README", index=False)

        wb = writer.book

        # Styling Helper
        header_fill = PatternFill("solid", fgColor="E6EEF8")
        bold = Font(bold=True)
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_sheet(name: str, freeze: bool = True, wide_wrap: bool = True):
            ws = wb[name]
            # Überschriften-Format
            if ws.max_row >= 1:
                for c in ws[1]:
                    c.font = bold
                    c.fill = header_fill
                    # eigenes Alignment-Objekt je Zelle
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = border

            # Inhalte
            if ws.max_row >= 2 and ws.max_column >= 1:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if wide_wrap:
                            try:
                                cell.alignment = cell.alignment.copy(horizontal="left", vertical="top", wrap_text=True)
                            except Exception:
                                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        cell.border = border

            # Spaltenbreiten (autofit grob, limitiert)
            for col_idx in range(1, ws.max_column + 1):
                col = get_column_letter(col_idx)
                maxlen = 8
                limit = min(ws.max_row, 200)  # Performance
                for r in range(1, limit + 1):
                    val = ws.cell(row=r, column=col_idx).value
                    if val is None:
                        continue
                    maxlen = max(maxlen, len(str(val)))
                ws.column_dimensions[col].width = min(maxlen + 2, 60)

            # Freeze Pane
            if freeze and ws.max_row > 1:
                ws.freeze_panes = "A2"
            return ws

        # Stil auf alle relevanten Blätter
        for sheet in ["01_Stammdaten", "10_Gefährdungen", "20_Maßnahmen", "30_Plan",
                      "40_Wirksamkeit", "50_Dokumentation", "60_Fortschreiben",
                      "90_Konfiguration", "99_README"]:
            wide = sheet not in ["01_Stammdaten", "90_Konfiguration", "99_README"]
            style_sheet(sheet, freeze=True, wide_wrap=wide)

        # Dropdown für Status im Plan-Blatt
        if "30_Plan" in wb.sheetnames:
            ws_plan = wb["30_Plan"]
            if ws_plan.max_row >= 2 and ws_plan.max_column >= 1:
                # Finde Spalte "Status"
                status_col_idx = None
                for c in range(1, ws_plan.max_column + 1):
                    if (ws_plan.cell(row=1, column=c).value or "").strip() == "Status":
                        status_col_idx = c
                        break
                if status_col_idx:
                    dv = DataValidation(
                        type="list",
                        formula1='"' + ",".join(STATUS_LIST) + '"',
                        allow_blank=True,
                        showDropDown=True,
                    )
                    ws_plan.add_data_validation(dv)
                    dv.add(f"{get_column_letter(status_col_idx)}2:{get_column_letter(status_col_idx)}1048576")

        # Farbskala (Risiko-Ampel) im Gefährdungsblatt auf "Risikosumme"
        if "10_Gefährdungen" in wb.sheetnames:
            ws_h = wb["10_Gefährdungen"]
            # Spalte "Risikosumme" suchen
            risk_col = None
            for c in range(1, ws_h.max_column + 1):
                if (ws_h.cell(row=1, column=c).value or "").strip() == "Risikosumme":
                    risk_col = c
                    break
            if risk_col:
                # 3-Farbskala: grün -> gelb -> rot
                col_letter = get_column_letter(risk_col)
                rng = f"{col_letter}2:{col_letter}{ws_h.max_row}"
                rule = ColorScaleRule(
                    start_type="num", start_value=1, start_color="C6EFCE",   # grünlich
                    mid_type="num", mid_value=max(2, thresholds[1]), mid_color="FFEB9C",  # gelb
                    end_type="num", end_value=max(3, thresholds[2] + 1), end_color="F8CBAD"  # rot
                )
                ws_h.conditional_formatting.add(rng, rule)

        # Druckfreundliche Kopfzeile (einfach)
        for name in wb.sheetnames:
            ws = wb[name]
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # beliebig viele Seiten in der Höhe

    bio.seek(0)
    return bio.read()





def as_json(assess: Assessment) -> str:
    return json.dumps(asdict(assess), ensure_ascii=False, indent=2)

def from_json(s: str) -> Assessment:
    data = json.loads(s)
    hazards = []
    for h in data.get("hazards", []):
        measures = [Measure(**m) for m in h.get("additional_measures", [])]
        hazards.append(Hazard(
            id=h["id"], area=h["area"], activity=h["activity"], hazard=h["hazard"],
            sources=h.get("sources", []),
            existing_controls=h.get("existing_controls", h.get("existing", [])),  # rückwärtskompatibel
            prob=h.get("prob", 3), sev=h.get("sev", 3),
            risk_value=h.get("risk_value", 9), risk_level=h.get("risk_level", "mittel"),
            additional_measures=measures, last_review=h.get("last_review"),
            reviewer=h.get("reviewer", ""), documentation_note=h.get("documentation_note", "")
        ))
    return Assessment(
        company=data.get("company",""), location=data.get("location",""),
        created_at=data.get("created_at",""), created_by=data.get("created_by",""),
        industry=data.get("industry","Hotel/Gastgewerbe"), scope_note=data.get("scope_note", ""),
        risk_matrix_thresholds=data.get("risk_matrix_thresholds", {"thresholds":[6,12,16]}),
        hazards=hazards, measures_plan_note=data.get("measures_plan_note",""),
        documentation_note=data.get("documentation_note",""), next_review_hint=data.get("next_review_hint","")
    )

def slug(*parts: str) -> str:
    s = "_".join(parts)
    s = re.sub(r"[^a-zA-Z0-9_-]+", "_", s)
    return s[:80]

# ===== Splitting für Mehrfach-Gefährdungen =====

_SPLIT_PATTERN = re.compile(r"\s*(?:,|/| und | & )\s*")

def split_hazard_text(text: str) -> List[str]:
    """Teilt 'Gefährdung' auf: Trennzeichen Komma, Slash, 'und', '&'."""
    if not text:
        return []
    parts = [p.strip() for p in _SPLIT_PATTERN.split(text) if p and p.strip()]
    # Duplikate eliminieren, Reihenfolge beibehalten
    seen, uniq = set(), []
    for p in parts:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq or [text.strip()]

# =========================
# Branchen-Bibliothek (ERWEITERT)
# =========================

def M(title, stop="O (Organisatorisch)"):
    return {"title": title, "stop_level": stop}

# --- HOTEL/GAST ---
LIB_HOTEL = {
    "Küche": [
        {"activity": "Kochen (Töpfe/Kessel)", "hazard": "Hitze, heiße Flüssigkeiten, Verbrühungen/Verbrennungen", "sources": ["Herde","Kessel","Töpfe"], "existing": ["Hitzeschutz"], "measures":[M("Topfdeckel/Spritzschutz nutzen","T (Technisch)"), M("‚Heiß!‘ rufen"), M("Hitzeschutzhandschuhe","P (PSA)")]},
        {"activity": "Braten (Pfanne/Grillplatte)", "hazard": "Fettspritzer, Verbrennungen, Rauch/Dämpfe", "sources": ["Pfannen","Grillplatten"], "existing": ["Abzug"], "measures":[M("Spritzschutz einsetzen","T (Technisch)"), M("Haube reinigen/prüfen")]},
        {"activity": "Frittieren", "hazard": "Fettbrand, Verbrennungen, Spritzer", "sources": ["Fritteusen"], "existing": ["Fettbrandlöscher"], "measures":[M("Ölwechsel-/Reinigungsplan"), M("Hitzeschutzschürze & Handschuhe","P (PSA)")]},
        {"activity": "Kombidämpfer öffnen", "hazard": "Dampf/Heißluft – Verbrühung beim Öffnen", "sources": ["Kombidämpfer"], "existing": ["Abkühlzeit"], "measures":[M("Tür erst spaltweise öffnen"), M("Hitzeschutzhandschuhe","P (PSA)")]},
        {"activity": "Saucen/Reduktionen", "hazard": "Dampf, Spritzer, inhalative Belastung", "sources": ["Reduktion"], "existing": ["Abluft"], "measures":[M("Deckel/Spritzschutz","T (Technisch)"), M("Lüftung checken")]},
        {"activity": "Schneiden mit Messern", "hazard": "Schnitt-/Stichverletzungen", "sources": ["Messer"], "existing": ["Scharfe Messer"], "measures":[M("Schleifplan"), M("Schnittschutzhandschuhe bei Bedarf","P (PSA)")]},
        {"activity": "Aufschnittmaschine", "hazard": "Schnittverletzungen an rotierenden Klingen", "sources": ["Aufschnitt"], "existing": ["Schutzhaube","Not-Aus"], "measures":[M("Sicherheitsbauteile prüfen","T (Technisch)"), M("Nur befugte Bedienung")]},
        {"activity": "Fleischwolf/Gemüseschneider", "hazard": "Eingezogenwerden, Schnittverletzung", "sources": ["Wolf","Gemüseschneider"], "existing": ["Stopfer"], "measures":[M("Stopfer verwenden"), M("Unterweisung Not-Aus","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Kippkessel/Bräter", "hazard": "Verbrühung, Quetschen beim Kippen", "sources": ["Kippkessel"], "existing": ["Hitzeschutz"], "measures":[M("Kipp-Prozess standardisieren"), M("Zweihandbedienung beachten","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Spülbereich", "hazard": "Heißes Wasser/Dampf, Chemikalien, Rutschgefahr", "sources": ["Spülmaschine","Klarspüler"], "existing": ["Hand-/Augenschutz"], "measures":[M("Sofort-Wisch-Regel"), M("Antirutsch-Matten","T (Technisch)")]},
        {"activity": "Reinigung/Chemie", "hazard": "Ätz-/Reizwirkung, Chlorgas bei Mischungen", "sources": ["Reiniger/Desinfektion"], "existing": ["Dosiersysteme"], "measures":[M("Vordosierte Kartuschen","S (Substitution/Quelle entfernen)"), M("Betriebsanweisungen aushängen")]},
        {"activity": "Gasgeräte", "hazard": "Gasleck, CO-Bildung, Brand/Explosion", "sources": ["Gasherde","Leitungen"], "existing": ["Dichtheitsprüfung"], "measures":[M("Gaswarnmelder","T (Technisch)"), M("Leckcheck vor Inbetriebnahme")]},
        {"activity": "Warenannahme/Hubwagen", "hazard": "Quetschungen, Heben/Tragen, Verkehrswege", "sources": ["Rollcontainer","Hubwagen"], "existing": ["Hebehilfen"], "measures":[M("Wege kennzeichnen"), M("Kurzunterweisung Heben/Tragen","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Altöl/Müll entsorgen", "hazard": "Verbrennung bei heißem Öl, Schnitt/Infektion", "sources": ["Altöl","Müllsack"], "existing": ["Abkühlen"], "measures":[M("Deckel-Transportbehälter","T (Technisch)"), M("Handschutz verpflichtend","P (PSA)")]},
        {"activity": "TK-/Kühlräume", "hazard": "Kälte, Rutschgefahr, Einsperr-Risiko", "sources": ["Kühlzelle","TK"], "existing": ["Kälteschutz"], "measures":[M("Tür-Notöffnung prüfen","T (Technisch)"), M("Aufenthaltsdauer begrenzen")]},
        {"activity": "Allergenmanagement", "hazard": "Kreuzkontamination/Allergene", "sources": ["Zutatenwechsel"], "existing": ["Kennzeichnung"], "measures":[M("Rein-/Unrein-Organisation"), M("Unterweisung LMIV","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Elektrische Kleingeräte", "hazard": "Stromschlag, Brandrisiko", "sources": ["Mixer","Pürierstab"], "existing": ["Sichtprüfung"], "measures":[M("Prüfintervall ortsveränderliche Geräte")]},
    ],
    "Housekeeping": [
        {"activity": "Betten machen", "hazard": "Rücken-/Schulterbelastung, Verdrehungen", "sources": ["Matratzen"], "existing": ["Arbeitstechnik"], "measures":[M("Ecken-Technik schulen","Q (Qualifikation/Unterweisung)"), M("Leichtere Bettwaren","S (Substitution/Quelle entfernen)")]},
        {"activity": "Sanitärreinigung", "hazard": "Chemikalienreizungen, Aerosole", "sources": ["Reiniger"], "existing": ["Hautschutzplan"], "measures":[M("Dosierstation/Piktogramme","T (Technisch)"), M("Sprühnebel vermeiden","S (Substitution/Quelle entfernen)")]},
        {"activity": "Fenster/Glas innen", "hazard": "Sturz, Schnitt an Glas", "sources": ["Leitern","Glas"], "existing": ["Leiterprüfung"], "measures":[M("Teleskopstiele statt Leiter","S (Substitution/Quelle entfernen)"), M("Schnittfeste Handschuhe","P (PSA)")]},
        {"activity": "Wäschetransport", "hazard": "Heben/Tragen, Quetschungen", "sources": ["Wäschewagen"], "existing": ["Schiebehilfen"], "measures":[M("Lastbegrenzung"), M("Türen offen sichern")]},
        {"activity": "Abfallentsorgung", "hazard": "Stich-/Schnittverletzungen, Infektionsgefahr", "sources": ["Scherben","Nadeln"], "existing": ["Feste Behälter"], "measures":[M("Sharps-Boxen","T (Technisch)"), M("Meldeweg Nadel-/Scherbenfund")]},
    ],
    "Service/Bar": [
        {"activity":"Heißgetränke zubereiten","hazard":"Verbrühungen/Verbrennungen","sources":["Kaffeemaschine"],"existing":["Hitzeschutz"],"measures":[M("Dampflanze abblasen"),M("Handschutz bereit","P (PSA)")]},
        {"activity":"Flambieren/Offene Flamme","hazard":"Brand/Alkoholdämpfe","sources":["Brenner","Spirituosen"],"existing":["Abstand"],"measures":[M("Nur geschultes Personal"),M("Löschmittel bereit")]},
        {"activity":"CO₂-Zapfanlage/Flaschenwechsel","hazard":"Erstickungsgefahr, Hochdruck","sources":["CO₂-Flaschen"],"existing":["CO₂-Warner"],"measures":[M("Sensorentest dokumentieren","T (Technisch)"),M("Wechsel nur zu zweit")]},
        {"activity":"Gläser polieren/Bruch","hazard":"Schnittverletzungen","sources":["Glas"],"existing":["Entsorgung"],"measures":[M("Polierhandschuhe","P (PSA)")]},
    ],
    "Technik/Haustechnik": [
        {"activity":"Elektroarbeiten (EUP/EFK)","hazard":"Elektrischer Schlag, Lichtbogen","sources":["Verteilungen"],"existing":["LOTO"],"measures":[M("LOTO-Verfahren dokumentieren"),M("PSA+Prüfer anwenden","T (Technisch)")]},
        {"activity":"Heißarbeiten (Schweißen/Trennen)","hazard":"Brand/Explosion, Rauch","sources":["Schweißgerät"],"existing":["Genehmigung","Feuerwache"],"measures":[M("Funkenschutz","T (Technisch)"),M("Nachkontrolle")]},
        {"activity":"Dach-/Höhenarbeit","hazard":"Absturz","sources":["Dachkanten"],"existing":["PSAgA"],"measures":[M("Anschlagpunkte prüfen","T (Technisch)"), "Rettungsplan"]},  # String absichtlich: wird robust verarbeitet
    ],
    "Lager/Wareneingang": [
        {"activity":"Auspacken/Öffnen","hazard":"Schnittverletzungen, Stolpern","sources":["Cutter","Umreifungen"],"existing":["Sichere Messer"],"measures":[M("Sicherheitsmesser einsetzen","S (Substitution/Quelle entfernen)"),M("Müll-Station nahe Rampe")]},
        {"activity":"Palettieren/Bewegen","hazard":"Quetschungen, Anfahren","sources":["Rollcontainer","Hubwagen"],"existing":["Wege markieren"],"measures":[M("Stopper an Rampen","T (Technisch)"),M("Verkehrsordnung aushängen")]},
        {"activity":"Hochregal/Entnahme in Höhe","hazard":"Absturz/Herabfallende Teile","sources":["Leitern","Regale"],"existing":["Leiterprüfung"],"measures":[M("Nur geprüfte Tritte"),M("Lastsicherung kontrollieren")]},
        {"activity":"TK-Lager/Kälte","hazard":"Kälte, Rutsch","sources":["Eis","Kondenswasser"],"existing":["Kälteschutz"],"measures":[M("Aufenthaltsdauer begrenzen"),M("Eis entfernen/Matten","T (Technisch)")]},
    ],
    "Spa/Wellness": [
        {"activity":"Sauna/Ofen & Aufguss","hazard":"Verbrennungen, Brand, Heißdampf","sources":["Saunaöfen"],"existing":["Abschirmungen"],"measures":[M("Ofenschutz/Temperaturwächter prüfen","T (Technisch)"),M("Aufgussregeln festlegen")]},
        {"activity":"Pooltechnik/Chemie","hazard":"Gefahrstoffe (Chlor, pH), Gasfreisetzung","sources":["Dosier-/Lagerräume"],"existing":["Lüftung/Absaugung"],"measures":[M("Auffangwannen/Trennung","T (Technisch)"),M("Freigabe mit Gaswarner")]},
        {"activity":"Nassbereiche","hazard":"Rutsch-/Sturzgefahr","sources":["Fliesen","Wasser"],"existing":["Rutschhemmung"],"measures":[M("Rutschmatten/Beläge prüfen","T (Technisch)"),M("Sofort-Wisch-Regel & Sperrung")]},
    ],
    "Rezeption": [
        {"activity":"Front Office/Gästekommunikation","hazard":"Psychische Belastung, Konflikte","sources":["Stoßzeiten"],"existing":["Deeskalation"],"measures":[M("Stoßzeiten doppelt besetzen")]},
        {"activity":"Nacht-/Alleinarbeit","hazard":"Überfall/Bedrohung, Ermüdung","sources":["Nachtschicht"],"existing":["Alarmtaster"],"measures":[M("Stillen Alarm testen","T (Technisch)"),M("Zwei-Personen-Regel nach Risiko")]},
        {"activity":"Bildschirm/Kasse","hazard":"Ergonomie, Augenbelastung","sources":["Monitore"],"existing":["Ergonomiecheck"],"measures":[M("20-20-20-Regel & Mikropausen"),M("Sehtest/Bildschirmbrille","Q (Qualifikation/Unterweisung)")]},
    ],
    "Verwaltung": [
        {"activity":"Bildschirmarbeit","hazard":"Haltungs-/Augenbelastung","sources":["Sitzplätze","Monitore"],"existing":["Höhenverstellbar"],"measures":[M("Monitorhöhe/Abstand einstellen","T (Technisch)"),"Mikropausenregelung"]},
        {"activity":"Laserdrucker/Toner","hazard":"Feinstaub, Hautkontakt","sources":["Tonerwechsel"],"existing":["Lüftung"],"measures":[M("Wechselhandschuhe/Abfallbeutel","T (Technisch)")]},
    ],
    "Außenbereiche": [
        {"activity":"Gartenpflege/Mähen","hazard":"Projektilwurf, Lärm","sources":["Rasenmäher"],"existing":["Schutzbrille","Gehörschutz"],"measures":[M("Steinkontrolle vor Start"),M("Visier/Gehörschutz","P (PSA)")]},
        {"activity":"Hecken-/Baumschnitt","hazard":"Schnittverletzung, Absturz","sources":["Heckenschere","Leiter"],"existing":["Leiter sichern"],"measures":[M("Teleskopgeräte statt Leiter","S (Substitution/Quelle entfernen)")]},
        {"activity":"Winterdienst","hazard":"Rutschen, Kälte","sources":["Eis/Schnee"],"existing":["Räum-/Streuplan"],"measures":[M("Rutschhemmende Spikes/Schuhe","P (PSA)"),M("Prioritätswege & Frühstartplan")]},
    ],
}
# --- Bäckerei (erweitert) ---
LIB_BAECKEREI = {
    "Rohstofflagerung/-handling": [
        {"activity":"Mehlsilo befüllen", "hazard":"Mehlstaubexposition", "sources":["Silo","Förderschnecke"], "existing":["Absaugung vorhanden"], "measures":[M("Absaugung prüfen/Protokoll","T (Technisch)"), M("FFP2/Staubmaske bereithalten","P (PSA)")]},
        {"activity":"Mehlsilo befüllen", "hazard":"Staubexplosionsgefahr", "sources":["Silo","Entlüftung"], "existing":["Erdung/Verdrahtung"], "measures":[M("Zündquellen vermeiden","O (Organisatorisch)"), M("ATEX-Doku bereithalten")]},
        {"activity":"Sackware tragen/heben", "hazard":"Überlastung Bewegungsapparat", "sources":["Mehlsäcke","Zutaten"], "existing":["Hubwagen"], "measures":[M("Hebe-/Tragetechnik schulen","Q (Qualifikation/Unterweisung)"), M("Lastbegrenzung einführen")]},
        {"activity":"Zutaten abwiegen (Nüsse, Milchpulver, Eier)", "hazard":"Allergenkontakt", "sources":["Allergene"], "existing":["Kennzeichnung"], "measures":[M("Rein-/Unrein-Trennung","O (Organisatorisch)"), M("Einmalhandschuhe bereitstellen","P (PSA)")]},
        {"activity":"TK-/Kühlraum betreten", "hazard":"Kälteexposition", "sources":["TK-Raum","Kühlraum"], "existing":["Kälteschutzjacken"], "measures":[M("Aufenthaltsdauer begrenzen"), M("Tür-Notöffnung prüfen","T (Technisch)")]},
        {"activity":"TK-/Kühlraum betreten", "hazard":"Rutschgefahr", "sources":["Kondenswasser","Glätte"], "existing":["Rutschhemmender Boden"], "measures":[M("Antirutsch-Matten einsetzen","T (Technisch)"), M("Sofort-Wisch-Regel einführen")]},
    ],

    "Teigherstellung": [
        {"activity":"Spiralkneter bedienen", "hazard":"Eingezogenwerden an bewegten Teilen", "sources":["Spiralkneter","Werkzeuge"], "existing":["Schutzhaube","Not-Aus"], "measures":[M("Interlocks prüfen/Journal","T (Technisch)"), M("Nur befugte Bedienung","O (Organisatorisch)")]},
        {"activity":"Hubkneter kippen", "hazard":"Quetschgefahr beim Kippen", "sources":["Hubkneter","Kippvorrichtung"], "existing":["Zweihandbedienung"], "measures":[M("Kippbereich freihalten","O (Organisatorisch)"), M("Jährliche Prüfung dokumentieren","T (Technisch)")]},
        {"activity":"Kippkneter reinigen", "hazard":"Unerwarteter Anlauf", "sources":["Antrieb","Schalter"], "existing":["Hauptschalter"], "measures":[M("LOTO/Stillsetzen anwenden","O (Organisatorisch)"), M("Unterweisung LOTO","Q (Qualifikation/Unterweisung)")]},
        {"activity":"Teigteiler bedienen", "hazard":"Quetsch-/Scherstellen", "sources":["Teigteiler"], "existing":["Schutzvorrichtungen"], "measures":[M("Sicherheitsfunktions-Test täglich","O (Organisatorisch)")]},
        {"activity":"Rundwirker bedienen", "hazard":"Fingerquetschung an Transportrollen", "sources":["Rundwirker"], "existing":["Abdeckungen"], "measures":[M("Gefahrenbereich markieren","T (Technisch)")]},
        {"activity":"Teig wiegen/portionieren", "hazard":"Ergonomische Zwangshaltungen", "sources":["Waage","Arbeitstische"], "existing":["Höhenverstellbarer Tisch"], "measures":[M("Arbeitsplätze anpassen","T (Technisch)"), M("Mikropausen planen","O (Organisatorisch)")]},
    ],

    "Backen": [
        {"activity":"Etagenofen beschicken", "hazard":"Verbrennung an heißen Flächen", "sources":["Etagenofen","Backkammer"], "existing":["Ofenhandschuhe"], "measures":[M("Hitzeschutzhandschuhe nutzen","P (PSA)"), M("Heißflächen kennzeichnen","T (Technisch)")]},
        {"activity":"Stikkenofen öffnen", "hazard":"Verbrühung durch Heißdampf", "sources":["Stikkenofen"], "existing":["Warnhinweise"], "measures":[M("Tür spaltweise öffnen","O (Organisatorisch)"), M("Gesicht abwenden")]},
        {"activity":"Schießerarbeit (ein-/ausbacken)", "hazard":"Überstreckung/Rückenbelastung", "sources":["Schießer","Bleche"], "existing":["Hilfsmittel"], "measures":[M("Schießer mit Stiel in passend","S (Substitution/Quelle entfernen)"), M("Zweitperson bei großen Blechen","O (Organisatorisch)")]},
        {"activity":"Backwagen fahren", "hazard":"Quetsch-/Anfahrgefahr", "sources":["Backwagen","Engstellen"], "existing":["Wegeordnung"], "measures":[M("Wege freihalten/markieren","T (Technisch)")]},
        {"activity":"Schmalzgebäck frittieren", "hazard":"Fettbrand", "sources":["Fritteuse","Fett"], "existing":["Fettbrandlöscher"], "measures":[M("Kein Wasser auf Fett!","O (Organisatorisch)"), M("Löschdecke bereithalten","P (PSA)")]},
        {"activity":"Schmalzgebäck frittieren", "hazard":"Verbrennung durch heißes Fett", "sources":["Fritteuse","Spritzer"], "existing":["Spritzschutz"], "measures":[M("Hitzeschutzschürze tragen","P (PSA)"), M("Ölstand/Temp. kontrollieren","O (Organisatorisch)")]},
    ],

    "Weiterverarbeitung": [
        {"activity":"Kuvertüre temperieren", "hazard":"Kontakt mit heißer Flüssigkeit", "sources":["Kuvertürebecken"], "existing":["Hitzeschutz"], "measures":[M("Spritzschutz verwenden","T (Technisch)"), M("Langsam umfüllen","O (Organisatorisch)")]},
        {"activity":"Füllmaschine betreiben", "hazard":"Quetschgefahr an Einzugspunkten", "sources":["Füllzylinder","Förderung"], "existing":["Abdeckungen"], "measures":[M("Nur bei Stillstand reinigen","O (Organisatorisch)"), M("Hauptschalter sichern (LOTO)")]},
        {"activity":"Kühl-/TK-Ware entnehmen", "hazard":"Kältebedingte Erfrierung", "sources":["TK-Regal","Minusgrade"], "existing":["Kälteschutzhandschuhe"], "measures":[M("Kälteschutz-PSA konsequent","P (PSA)"), M("Kurzzeit-Aufenthalte planen","O (Organisatorisch)")]},
        {"activity":"Torten/Feingebäck glasieren", "hazard":"Rutschgefahr durch Tropfen", "sources":["Glasur","Boden"], "existing":["Bodenreinigung"], "measures":[M("Anti-Rutsch-Matten","T (Technisch)"), M("Sofort-Wisch-Regel","O (Organisatorisch)")]},
    ],

    "Reinigung": [
        {"activity":"Maschinenreinigung (Teigteiler/Rundwirker)", "hazard":"Schnitt-/Schergefahr an Klingen", "sources":["Messer","Klingen"], "existing":["Werkzeug zum Entfernen"], "measures":[M("Reinigung nur stromlos","O (Organisatorisch)"), M("Schnittschutzhandschuhe","P (PSA)")]},
        {"activity":"CIP/Schlauchreinigung", "hazard":"Ätz-/Reizwirkung durch Chemie", "sources":["Lauge","Säure"], "existing":["Dosieranlage"], "measures":[M("Betriebsanweisungen aushängen","O (Organisatorisch)"), M("Augendusche/Notdusche prüfen","T (Technisch)")]},
        {"activity":"Bodenreinigung Nassbereich", "hazard":"Rutsch-/Sturzgefahr", "sources":["Wasser","Reiniger"], "existing":["Warnschilder"], "measures":[M("Abschnittweise sperren","O (Organisatorisch)"), M("Rutschhemmende Schuhe","P (PSA)")]},
        {"activity":"Backöfen reinigen", "hazard":"Verbrennung durch Restwärme", "sources":["Backraum","Heizelemente"], "existing":["Abkühlzeit"], "measures":[M("Freigabe erst <40°C","O (Organisatorisch)"), M("Hitzeschutz-Handschuhe","P (PSA)")]},
    ],

    "Verkauf/Service": [
        {"activity":"Brotschneidemaschine bedienen", "hazard":"Schnittverletzung an der Klinge", "sources":["Brotschneider"], "existing":["Schutzhaube"], "measures":[M("Nur befugte Bedienung","O (Organisatorisch)"), M("Schnittschutzhandschuhe bereit","P (PSA)")]},
        {"activity":"Heißgetränke ausgeben", "hazard":"Verbrühung an heißem Wasser/Dampf", "sources":["Kaffeemaschine","Heißwasser"], "existing":["Warnhinweis"], "measures":[M("Dampflanze vorher abblasen","O (Organisatorisch)"), M("Becher sicher abstellen")]},
        {"activity":"Verkaufstheke bedienen", "hazard":"Infektionsrisiko durch Kundenkontakt", "sources":["Publikumsverkehr"], "existing":["Händehygiene"], "measures":[M("Händedesinfektion bereitstellen","T (Technisch)")]},
        {"activity":"Kassentätigkeit", "hazard":"Psychische Belastung/Überfallrisiko", "sources":["Kasse","Kundenstrom"], "existing":["Kassenrichtlinie"], "measures":[M("Deeskalation schulen","Q (Qualifikation/Unterweisung)"), M("Keine großen Barbestände","O (Organisatorisch)")]},
    ],

    "Logistik/Transport": [
        {"activity":"Backwagen/Rollwagen schieben", "hazard":"Hand-/Fingerquetschung", "sources":["Backwagen","Türstock"], "existing":["Griffe"], "measures":[M("Langsam in Engstellen","O (Organisatorisch)"), M("Anschlagkanten schützen","T (Technisch)")]},
        {"activity":"Rampe/Andienung", "hazard":"Stolper-/Sturzgefahr an Kanten", "sources":["Rampe","Kanten"], "existing":["Kantenmarkierung"], "measures":[M("Rampe sauber/vereisungsfrei","O (Organisatorisch)")]},
        {"activity":"Beladen von Lieferfahrzeugen", "hazard":"Überlastung/Heben/Tragen", "sources":["Kisten","Bleche"], "existing":["Rollhilfen"], "measures":[M("Team-Hebeplan/Lastbegrenzung","O (Organisatorisch)"), M("Hebehilfen beschaffen","T (Technisch)")]},
        {"activity":"Stapler/FFZ im Lager", "hazard":"Anfahr-/Quetschunfälle", "sources":["Stapler","Hubwagen"], "existing":["Wegeordnung"], "measures":[M("Fuß-/Fahrwege trennen","T (Technisch)"), M("Nur ausgebildete Fahrer","O (Organisatorisch)")]},
    ],

    "Wartung/Technik": [
        {"activity":"Elektrische Instandhaltung", "hazard":"Elektrischer Schlag", "sources":["Verteilungen","Maschinen"], "existing":["E-Check"], "measures":[M("Nur EFK/EUP beauftragen","O (Organisatorisch)"), M("Spannungsfreiheit prüfen","O (Organisatorisch)")]},
        {"activity":"Gasbrenner/Öfen warten", "hazard":"Explosions-/Brandgefahr", "sources":["Gasleitungen","Brenner"], "existing":["Dichtheitsprüfung"], "measures":[M("Lecksuche/Seifentest","O (Organisatorisch)"), M("Feuerlöscher bereitstellen")]},
        {"activity":"Druckluftanlage betreiben", "hazard":"Schlauchplatzer/Lärm", "sources":["Kompressor","Schläuche"], "existing":["Wartungsplan"], "measures":[M("Geprüfte Schläuche/Armaturen","T (Technisch)"), M("Gehörschutz verfügbar","P (PSA)")]},
        {"activity":"Kälteanlage (NH3/CO₂)", "hazard":"Gefahrstofffreisetzung", "sources":["Kälteanlage","Aggregate"], "existing":["Gaswarnanlage"], "measures":[M("Alarm-/Rettungsplan aushängen","O (Organisatorisch)"), M("Fluchtfilter/Fluchtgeräte","P (PSA)")]},
    ],
}


# --- Fleischerei/Metzgerei ---
LIB_FLEISCHEREI = {
    "Produktion": [
        {"activity":"Bandsäge","hazard":"Schnitt/Amputation","sources":["Bandsäge"],"existing":["Schutzhaube","Not-Aus"],"measures":[M("Nur befugte Bedienung"),M("Reinigung stromlos")]},
        {"activity":"Fleischwolf","hazard":"Eingezogenwerden","sources":["Fleischwolf"],"existing":["Stopfer","Schutz"],"measures":[M("Stopfer konsequent nutzen")]},
        {"activity":"Kutter","hazard":"Schnitt/Schlag","sources":["Kutter"],"existing":["Haube","Verriegelung"],"measures":[M("Verriegelung prüfen","T (Technisch)")]},
        {"activity":"Vakuumierer/Schrumpfer","hazard":"Verbrennung/Quetschung","sources":["Heißsiegel"],"existing":["Hitzeschutz"],"measures":[M("Heißzonen markieren","T (Technisch)")]},
        {"activity":"Kühl-/TK-Lager","hazard":"Kälte/Rutsch","sources":["Kühl/TK"],"existing":["Kälteschutz"],"measures":[M("Zeitbegrenzung/Matten")]},
        {"activity":"Reinigung/Desinfektion","hazard":"Chemische Belastung","sources":["Reiniger"],"existing":["PSA"],"measures":[M("SDB/Betriebsanweisungen an Station","T (Technisch)")]},
        {"activity":"Räuchern/Heißräuchern","hazard":"Rauch/Verbrennung/CO","sources":["Räucherkammer"],"existing":["Abluft"],"measures":[M("CO-Warnung falls nötig","T (Technisch)")]},
    ],
    "Verkauf": [
        {"activity":"Aufschnitt/Bedienung","hazard":"Schnittverletzung","sources":["Aufschnitt"],"existing":["Schutzhaube"],"measures":[M("Messerschulung/Handschutz","Q (Qualifikation/Unterweisung)")]},
        {"activity":"Heißtheke","hazard":"Verbrennung","sources":["Heiße Theken"],"existing":["Hitzeschutz"],"measures":[M("Abdeckung/Abstellen sichern","T (Technisch)")]},
    ]
}

# --- Gemeinschaftsverpflegung/Kantine ---
LIB_KANTINE = {
    "Küche": [
        {"activity":"Großkochgeräte/Kippkessel","hazard":"Verbrühung, Quetschung beim Kippen","sources":["Kippkessel"],"existing":["Hitzeschutz","2-Hand-Bed."],"measures":[M("Kipp-Prozess standardisieren")]},
        {"activity":"Tablettförderband/Spülstraße","hazard":"Einklemm-/Scherstellen, Heißwasser/Dampf","sources":["Bandspülmaschine"],"existing":["Abdeckungen","Not-Aus"],"measures":[M("Nur befugte Bedienung")]},
        {"activity":"Ausgabe/Frontcooking","hazard":"Verbrennung, Kontakt mit Gästen","sources":["Wärmebrücken","Pfannen"],"existing":["Abschirmung","Greifzonen"],"measures":[M("Abstand/Abschirmung","T (Technisch)")]},
        {"activity":"Regenerieren/Heißluftwagen","hazard":"Verbrennung, Dampf","sources":["Heißluftwagen"],"existing":["Hitzeschutz"],"measures":[M("Türöffnungsroutine"),M("Schutzhandschuhe","P (PSA)")]},
    ],
    "Logistik": [
        {"activity":"Transportwagen/Tablettwagen","hazard":"Quetschen/Stolpern","sources":["Rollwagen","Aufzüge"],"existing":["Wege frei"],"measures":[M("Lastbegrenzung/Wegepriorität")]},
        {"activity":"Annahme/Kommissionierung","hazard":"Schnitt/Heben/Tragen","sources":["Kisten","Folien"],"existing":["Sichere Messer","Rollwagen"],"measures":[M("Sicherheitsmesser einsetzen","S (Substitution/Quelle entfernen)")]},
    ]
}

# --- Konditorei/Café ---
LIB_KONDITOREI = {
    "Produktion": [
        {"activity":"Zucker kochen/Karamell","hazard":"Heißsirup/Verbrennung","sources":["Kocher"],"existing":["Hitzeschutz"],"measures":[M("Schutzbrille & langsames Aufgießen","P (PSA)")]},
        {"activity":"Kuvertüre/Temperieren","hazard":"Hitze, Spritzer","sources":["Bad/Tempering"],"existing":["Hitzeschutz"],"measures":[M("Deckel/Spritzschutz","T (Technisch)")]},
        {"activity":"Kleingeräte/Rührwerke","hazard":"Scher-/Einklemmstellen","sources":["Rührwerk"],"existing":["Schutz","Not-Aus"],"measures":[M("Nur stromlos reinigen")]},
        {"activity":"Kühl-/TK","hazard":"Kälte/Rutsch","sources":["Kühl/TK"],"existing":["Kälteschutz"],"measures":[M("Aufenthalt begrenzen/Eis entfernen")]},
        {"activity":"Reinigung","hazard":"Chemikalien","sources":["Reiniger"],"existing":["PSA"],"measures":[M("Dosierhilfen/Betriebsanweisung","T (Technisch)")]},
    ],
    "Verkauf/Café": [
        {"activity":"Kaffeemaschine/Heißgetränke","hazard":"Verbrühung","sources":["Dampflanze"],"existing":["Hitzeschutz"],"measures":[M("Dampflanze abblasen")]},
        {"activity":"Tortenmesser/Glasvitrine","hazard":"Schnitt/Glasschaden","sources":["Glas","Messer"],"existing":["Sichere Entsorgung"],"measures":[M("Polier-/Schnittschutzhandschuhe","P (PSA)")]},
    ]
}

# --- Brauerei ---
LIB_BRAUEREI = {
    "Sudhaus": [
        {"activity":"Maischen/Kochen im Sudkessel","hazard":"Heißdampf/Verbrühung, CO₂ beim Kochen","sources":["Sudkessel","Whirlpool"],"existing":["Abschrankung","Hitzeschutz"],"measures":[M("Deckel & Dampfableitung prüfen","T (Technisch)"),M("Heißarbeiten vermeiden, Vorsicht beim Öffnen")]},
        {"activity":"Whirlpool/Trubabzug","hazard":"Heißdampf/Verbrennung","sources":["Whirlpool"],"existing":["Abdeckung"],"measures":[M("Öffnen nur nach Abkühlen")]},
        {"activity":"Läuterbottich","hazard":"Einsinken/Erstickung bei Einstieg, Heißdampf","sources":["Läuterbottich"],"existing":["Zutritt verboten"],"measures":[M("Befahren als enge Räume regeln (Permit)")]},
        {"activity":"Reinigung CIP","hazard":"Ätz-/Reizwirkung, Gasbildung","sources":["Laugen/Säuren"],"existing":["Dosierung","BA"],"measures":[M("CIP-Schläuche sichern","T (Technisch)"),M("Augendusche/Notdusche prüfen","T (Technisch)")]},
    ],
    "Gär-/Keller": [
        {"activity":"Gär-/Lagertanks","hazard":"CO₂-Ansammlung/Erstickung, Druck","sources":["Gärtank"],"existing":["CO₂-Warner","Lüftung"],"measures":[M("Warner testen & loggen","T (Technisch)"),M("Freimessen vor Einstieg")]},
        {"activity":"Druckbehälter/Überdruck","hazard":"Explosion/Druckverletzung","sources":["Tankdruck"],"existing":["Sicherheitsventile"],"measures":[M("SV-Prüfungen dokumentieren")]},
        {"activity":"Hefeernte/Umfüllen","hazard":"Biologische Gefährdung, Rutsch","sources":["Hefeschlamm"],"existing":["Handschutz"],"measures":[M("Spritzschutz & Kennzeichnung","T (Technisch)")]},
    ],
    "Abfüllung/Fasskeller": [
        {"activity":"Fassreinigung/Spülen","hazard":"CO₂/Restdruck, Chemie","sources":["Fasskeller"],"existing":["Druckentlastung"],"measures":[M("Entlüften/Spülen dokumentieren")]},
        {"activity":"Fassfüllen/Anstechen","hazard":"Druck, Schläge","sources":["Fass","ZKG"],"existing":["Sichere Kupplungen"],"measures":[M("Schlagschutz/PSA","P (PSA)")]},
    ],
    "Wartung/Technik": [
        {"activity":"CO₂-Flaschenlager","hazard":"Erstickung bei Leck","sources":["Flaschenbündel"],"existing":["CO₂-Warner","Belüftung"],"measures":[M("Dichtheitskontrolle")]},
        {"activity":"Ammoniak-Kälte","hazard":"NH₃-Toxizität/Leck","sources":["Kälteanlage"],"existing":["Gaswarnanlage"],"measures":[M("Alarm-/Rettungsplan"),M("Filter/Fluchtgeräte","P (PSA)")]},
    ],
}

# --- Getränkeabfüllung ---
LIB_GETRAENKEABF = {
    "Sirupe/Konzentrat": [
        {"activity":"Ansatz Sirup","hazard":"Chemische Reizung (Säuren/Basen), Rutsch","sources":["Zutaten","CIP"],"existing":["Dosierhilfen"],"measures":[M("BA & SDB an Station","T (Technisch)")]},
        {"activity":"Zuckerhandling","hazard":"Staubexplosion (selten), Ergonomie","sources":["Zucker"],"existing":["Absaugung"],"measures":[M("Staubarme Beschickung","S (Substitution/Quelle entfernen)")]},
    ],
    "Gebindehandling": [
        {"activity":"Leergutannahme/Sortierung","hazard":"Scherben/Schnitt, Lärm","sources":["Kästen","Flaschen"],"existing":["Handschutz","Gehörschutz"],"measures":[M("Scherbenbeseitigung sofort"),M("Lärmmonitoring")]},
        {"activity":"Waschmaschine","hazard":"Heißlauge, Dampf","sources":["Flaschenwascher"],"existing":["Einhausung"],"measures":[M("Spritzschutz & Handschutz","P (PSA)")]},
    ],
    "Füller/Etikettierer": [
        {"activity":"Füllerbereich","hazard":"Quetschen, Drehteile, Reinigungschemie","sources":["Füller","Transportbänder"],"existing":["Schutzzäune","Lichtgitter"],"measures":[M("Interlocks prüfen","T (Technisch)")]},
        {"activity":"CO₂-/Kohlensäureversorgung","hazard":"Erstickung, Hochdruck","sources":["CO₂-Tank"],"existing":["CO₂-Warner"],"measures":[M("Umfeld lüften, Sensorcheck","T (Technisch)")]},
    ],
    "Palettierung/Logistik": [
        {"activity":"Packen/Palettierer","hazard":"Einklemm-/Quetschstellen","sources":["Palettierer","Stretch"],"existing":["Schutzzonen"],"measures":[M("Sperrkreis & Freigabeprozesse")]},
        {"activity":"Flurförderzeuge","hazard":"Anfahren/Kollision","sources":["Stapler","Ameise"],"existing":["Wegeordnung"],"measures":[M("Staplerschein/Unterweisung","Q (Qualifikation/Unterweisung)")]},
    ]
}

# --- Eisherstellung ---
LIB_EIS = {
    "Produktion": [
        {"activity":"Pasteurisieren Milchmischung","hazard":"Verbrühung, Dampf","sources":["Pasteur"],"existing":["Hitzeschutz"],"measures":[M("Temperatur/Zeiten protokollieren")]},
        {"activity":"Homogenisieren/Mischen","hazard":"Einklemm-/Scherstellen","sources":["Homogenisator","Rührwerk"],"existing":["Schutzhauben"],"measures":[M("Reinigung nur stromlos")]},
        {"activity":"Gefrieren/Freezer","hazard":"Kälte/Erfrierung, Bewegte Teile","sources":["Kontifreezer"],"existing":["Abdeckungen"],"measures":[M("PSA Kälteschutz","P (PSA)")]},
        {"activity":"Aromen/Allergene","hazard":"Allergische Reaktionen/Kreuzkontamination","sources":["Nüsse","Milch"],"existing":["Allergenplan"],"measures":[M("Rein-/Unrein-Trennung")]},
        {"activity":"CIP-Reinigung","hazard":"Säuren/Laugen","sources":["CIP"],"existing":["Dosierung"],"measures":[M("Augendusche/Notdusche","T (Technisch)")]},
    ],
    "Verkauf/Theke": [
        {"activity":"Eistheke/Spatel","hazard":"Biologische Risiken, Temperaturkette","sources":["Theke"],"existing":["Temperaturkontrolle"],"measures":[M("Stichproben/Protokoll")]},
        {"activity":"Waffeleisen/Heißgeräte","hazard":"Verbrennung","sources":["Waffeleisen"],"existing":["Hitzeschutz"],"measures":[M("Handschutz bereit","P (PSA)")]},
    ],
    "Lager": [
        {"activity":"TK-Lager -30°C","hazard":"Kälte, Rutsch","sources":["TK"],"existing":["Kälteschutz"],"measures":[M("Max. Aufenthaltsdauer/Partnerprinzip")]},
    ]
}

# --- Event/Catering ---
LIB_EVENT = {
    "Vorbereitung/Produktion": [
        {"activity":"Mise en place/Kochen vor Ort","hazard":"Verbrennung/Verbrühung, Elektrik mobil","sources":["Induktionsfelder","Gasbrenner"],"existing":["E-Check mobil"],"measures":[M("Zuleitungen sichern"),M("Feuerlöscher bereit")]},
        {"activity":"Verladen/Transport","hazard":"Quetschung/Heben/Tragen","sources":["Kisten","GN-Behälter"],"existing":["Rollwagen"],"measures":[M("Ladungssicherung")]},
    ],
    "Aufbau/Betrieb": [
        {"activity":"Zelte/Provisorien","hazard":"Wind/Absturz/Stolpern","sources":["Zelt","Kabel"],"existing":["Abspannung","Kabelbrücken"],"measures":[M("Abnahme/Prüfbuch Zelt/Aggregat")]},
        {"activity":"Stromerzeuger/Aggregate","hazard":"CO/Abgase, Lärm, Stromschlag","sources":["Generator"],"existing":["Abstand/Lüftung"],"measures":[M("Erdung/PRCD-S","T (Technisch)"),M("CO-Warnung in Gebäuden","T (Technisch)")]},
        {"activity":"Ausgabe/Frontcooking","hazard":"Kontakt Gäste, heiße Flächen","sources":["Rechauds","Pfannen"],"existing":["Abschirmung"],"measures":[M("Greifzonen/Barriere","T (Technisch)")]},
    ],
    "Abbau/Reinigung": [
        {"activity":"Heißgeräte abbauen","hazard":"Verbrennung/Restwärme","sources":["Geräte"],"existing":["Abkühlen"],"measures":[M("Schnittschutzhandschuhe beim Packen","P (PSA)")]},
    ]
}

# --- Fast Food / Quickservice ---
LIB_QSR = {
    "Küche": [
        {"activity":"Fritteusenbetrieb","hazard":"Fettbrand, Verbrennung","sources":["Fritteuse"],"existing":["Löschdecke"],"measures":[M("Autom. Löschanlage (falls vorhanden) prüfen","T (Technisch)"),M("Kein Wasser!")]},
        {"activity":"Griddle/Flame Broiler","hazard":"Hitze/Verbrennung, Rauch","sources":["Grill"],"existing":["Abzug"],"measures":[M("Reinigungsplan Haube/Filter")]},
        {"activity":"Slicer/Chopper","hazard":"Schnitt/Scherstellen","sources":["Slicer"],"existing":["Schutz"],"measures":[M("Nur mit Werkzeug reinigen")]},
        {"activity":"Gefriertruhe/Schockfroster","hazard":"Kälte/Rutsch","sources":["TK"],"existing":["Kälteschutz"],"measures":[M("Eis entfernen")]},
        {"activity":"Bestellung/Allergene","hazard":"Fehlbestellung/Allergischer Schock","sources":["Kasse","App"],"existing":["Allergenliste"],"measures":[M("Abfrage Allergene im Bestellprozess")]},
    ],
    "Service": [
        {"activity":"Drive-Thru","hazard":"Fahrzeugkontakt/Abgase/Lärm","sources":["Fahrspur"],"existing":["Markierung"],"measures":[M("Sichtbarkeit/Reflexwesten","P (PSA)")]},
        {"activity":"Getränkespender/CO₂","hazard":"Erstickung/Hochdruck","sources":["CO₂-Flaschen"],"existing":["Befestigung"],"measures":[M("Sensorentest/Wechselprozess")]},
    ],
    "Reinigung": [
        {"activity":"Schaum-/Sprühreinigung","hazard":"Aerosole/Chemie","sources":["Reiniger"],"existing":["PSA"],"measures":[M("Schaumlanze statt Spray","S (Substitution/Quelle entfernen)")]},
    ]
}

# --- Wäscherei / Textilreinigung ---
LIB_WAESCHE = {
    "Annahme/Vorsortierung": [
        {"activity":"Schmutzwäscheannahme","hazard":"Biologische Gefährdungen, Stichverletzung","sources":["Schmutzwäsche"],"existing":["Handschutz"],"measures":[M("Sharps-Check/Trennung Unrein/Rein")]},
        {"activity":"Sortieren/Wiegen","hazard":"Heben/Tragen/Staub","sources":["Säcke","Wäschewagen"],"existing":["Hebehilfen"],"measures":[M("Absaugung an Entleerer","T (Technisch)")]},
    ],
    "Waschen/Nassreinigung": [
        {"activity":"Maschinenbeschickung","hazard":"Einklemm-/Scherstellen, Heißwasser/Dampf","sources":["Waschmaschinen"],"existing":["Not-Aus"],"measures":[M("Türverriegelungen prüfen","T (Technisch)")]},
        {"activity":"Chemiedosierung","hazard":"Ätz-/Reizwirkung","sources":["Flüssigchemie"],"existing":["Dosieranlage"],"measures":[M("Schlauch-/Kopplungscheck")]},
    ],
    "Finish/Trocknen/Mangeln": [
        {"activity":"Trockner/Mangel","hazard":"Einzugs-/Quetschstellen, Hitze","sources":["Tumbler","Mangel"],"existing":["Hauben","Zweihand"],"measures":[M("Einzugsabstand/Notleinen prüfen","T (Technisch)")]},
        {"activity":"Bügeln/Dampf","hazard":"Verbrühung/Verbrennung","sources":["Dampfbügel"],"existing":["Hitzeschutz"],"measures":[M("Dampfschläuche prüfen")]},
    ],
    "Reparatur/Nähen": [
        {"activity":"Nähmaschinenarbeit","hazard":"Nadelstich/Ergonomie","sources":["Nähmaschine"],"existing":["Fingerschutz"],"measures":[M("Beleuchtung/Arbeitshöhe anpassen","T (Technisch)")]},
    ],
}

INDUSTRY_LIBRARY: Dict[str, Dict[str, List[Dict[str, Any]]]] = {
    "Hotel/Gastgewerbe": LIB_HOTEL,
    "Bäckerei": LIB_BAECKEREI,
    "Fleischerei/Metzgerei": LIB_FLEISCHEREI,
    "Gemeinschaftsverpflegung/Kantine": LIB_KANTINE,
    "Konditorei/Café": LIB_KONDITOREI,
    "Brauerei": LIB_BRAUEREI,
    "Getränkeabfüllung": LIB_GETRAENKEABF,
    "Eisherstellung": LIB_EIS,
    "Event/Catering": LIB_EVENT,
    "Fast Food/Quickservice": LIB_QSR,
    "Wäscherei/Textilreinigung": LIB_WAESCHE,
}

# =========================
# Vorlagen laden/auswählen
# =========================

def add_template_items(
    assess: Assessment,
    template: Dict[str, List[Dict[str, Any]]],
    selected_keys: Optional[List[str]] = None,
    industry_name: Optional[str] = None,
    split_multi: Optional[bool] = None
):
    """Fügt Items aus einer Branchenvorlage hinzu.
    Robust: akzeptiert measures sowohl als Dicts (via M(...)) als auch als Strings.
    Optional: Multi-Gefährdungen splitten.
    """
    if split_multi is None:
        split_multi = st.session_state.get("opt_split_multi_hazards", True)

    DEFAULT_STOP = "O (Organisatorisch)"

    def normalize_measure(m: Any) -> Optional[Measure]:
        if isinstance(m, dict):
            return Measure(
                title=(m.get("title") or "").strip(),
                stop_level=m.get("stop_level", DEFAULT_STOP),
                notes=m.get("notes", "")
            )
        elif isinstance(m, str):
            t = m.strip()
            return Measure(title=t, stop_level=DEFAULT_STOP) if t else None
        else:
            return None

    for area, items in template.items():
        for item in items:
            key = template_item_key(industry_name or assess.industry, area, item)
            if selected_keys is not None and key not in selected_keys:
                continue

            hazard_text = item.get("hazard", "")
            hazards_list = split_hazard_text(hazard_text) if split_multi else [hazard_text]

            for hz_text in hazards_list:
                hz = Hazard(
                    id=new_id(),
                    area=area,
                    activity=item.get("activity", ""),
                    hazard=hz_text,
                    sources=item.get("sources", []) or [],
                    existing_controls=item.get("existing", []) or []
                )
                for m in item.get("measures", []) or []:
                    mm = normalize_measure(m)
                    if mm and mm.title:
                        hz.additional_measures.append(mm)
                assess.hazards.append(hz)

def preload_industry(assess: Assessment, industry_name: str, replace: bool = True):
    assess.industry = industry_name
    if replace:
        assess.hazards = []
    template = INDUSTRY_LIBRARY.get(industry_name, {})
    add_template_items(assess, template, selected_keys=None, industry_name=industry_name)

def template_item_key(industry: str, area: str, item: Dict[str, Any]) -> str:
    return slug(industry, area, item.get("activity",""), item.get("hazard",""))

def iter_template_items(industry: str) -> List[Tuple[str, Dict[str, Any], str]]:
    lib = INDUSTRY_LIBRARY.get(industry, {})
    out = []
    for area, items in lib.items():
        for it in items:
            out.append((area, it, template_item_key(industry, area, it)))
    return out

# =========================
# Streamlit App
# =========================

st.set_page_config(page_title="Gefährdungsbeurteilung – Branchen (BGN) mit Auswahl", layout="wide")

# Session init
if "assessment" not in st.session_state or st.session_state.get("assessment") is None:
    st.session_state.assessment = Assessment(
        company="Musterbetrieb GmbH", location="Beispielstadt",
        created_at=date.today().isoformat(), created_by="HSE/SiFa",
        industry="Hotel/Gastgewerbe",
    )
    preload_industry(st.session_state.assessment, "Hotel/Gastgewerbe", replace=True)

assess: Assessment = st.session_state.assessment

# Kopf
col_head1, col_head2 = st.columns([0.8, 0.2])
with col_head1:
    st.title("Gefährdungsbeurteilung – Branchen (BGN) mit Checkbox-Auswahl")
with col_head2:
    if st.button("📄 Duplizieren", key="btn_duplicate"):
        assess.created_at = date.today().isoformat()
        assess.company = f"{assess.company} (Kopie)"
        st.success("Kopie erstellt. Bitte speichern/exportieren.")

st.caption("Struktur: Vorlagen auswählen → Vorbereiten → Ermitteln → Beurteilen → Maßnahmen → Umsetzen → Wirksamkeit → Dokumentieren → Fortschreiben")

# Sidebar
with st.sidebar:
    st.header("Stammdaten")
    assess.company = st.text_input("Unternehmen", assess.company, key="meta_company")
    assess.location = st.text_input("Standort", assess.location, key="meta_location")
    assess.created_by = st.text_input("Erstellt von", assess.created_by, key="meta_created_by")
    assess.created_at = st.text_input("Erstellt am (ISO)", assess.created_at, key="meta_created_at")

    st.markdown("---")
    st.subheader("Branche wählen (für Vorlagen)")
    options = list(INDUSTRY_LIBRARY.keys())
    current_industry = getattr(assess, "industry", None) or "Hotel/Gastgewerbe"
    default_idx = options.index(current_industry) if current_industry in options else 0
    sector = st.selectbox("Branche", options=options, index=default_idx, key="sel_industry")
    st.caption(f"Aktuell geladen: **{assess.industry}**")

    # --- Optionen ---
    st.markdown("---")
    st.subheader("Optionen")
    if "opt_split_multi_hazards" not in st.session_state:
        st.session_state["opt_split_multi_hazards"] = True
    st.checkbox(
        "Mehrfach-Gefährdungen einer Tätigkeit automatisch auftrennen (1 Tätigkeit → 1 Gefährdung pro Eintrag)",
        key="opt_split_multi_hazards",
    )

    # Optional: Automatisches Nachladen bei Branchenwechsel
    st.markdown("---")
    st.caption("Automatisches Laden beim Branchenwechsel (optional)")
    if "last_sector" not in st.session_state:
        st.session_state.last_sector = sector
    elif st.session_state.last_sector != sector:
        assess.hazards = []
        tmpl = INDUSTRY_LIBRARY.get(sector, {})
        add_template_items(assess, tmpl, selected_keys=None, industry_name=sector)
        assess.industry = sector
        st.session_state.last_sector = sector
        st.toast(f"Vorlage '{sector}' automatisch geladen.", icon="✅")
        st.rerun()

    # --- Schnell-Laden der Branchenvorlage in der Sidebar ---
    st.markdown("---")
    st.markdown("**Schnell laden:**")
    c_load1, c_load2 = st.columns(2)
    with c_load1:
        if st.button("📚 Vorlage ERSETZEN", key="btn_load_replace_sidebar"):
            assess.hazards = []
            tmpl = INDUSTRY_LIBRARY.get(sector, {})
            add_template_items(assess, tmpl, selected_keys=None, industry_name=sector)
            assess.industry = sector
            if "template_checks" in st.session_state:
                st.session_state.template_checks = {}
            st.success(f"Vorlage '{sector}' geladen (ersetzt).")
            st.rerun()
    with c_load2:
        if st.button("➕ Vorlage ANHÄNGEN", key="btn_load_append_sidebar"):
            tmpl = INDUSTRY_LIBRARY.get(sector, {})
            add_template_items(assess, tmpl, selected_keys=None, industry_name=sector)
            assess.industry = sector
            st.success(f"Vorlage '{sector}' hinzugefügt (angehängt).")
            st.rerun()

    st.markdown("---")
    st.subheader("Risikomatrix (5×5)")
    thr = assess.risk_matrix_thresholds.get("thresholds", [6, 12, 16])
    low = st.number_input("Grenze niedrig (≤)", min_value=2, max_value=10, value=int(thr[0]), key="thr_low")
    mid = st.number_input("Grenze mittel (≤)", min_value=low+1, max_value=16, value=int(thr[1]), key="thr_mid")
    high = st.number_input("Grenze hoch (≤)", min_value=mid+1, max_value=24, value=int(thr[2]), key="thr_high")
    assess.risk_matrix_thresholds["thresholds"] = [low, mid, high]

    st.markdown("---")
    st.subheader("Export / Speicher")
    if st.button("📥 JSON sichern (Download unten aktualisieren)", key="btn_json_dump"):
        st.session_state["json_blob"] = as_json(assess)
    json_blob = st.session_state.get("json_blob", as_json(assess))
    st.download_button("⬇️ Download JSON", data=json_blob, file_name="gefaehrdungsbeurteilung.json", mime="application/json", key="btn_dl_json")

    excel_bytes = dump_excel(assess)
    st.download_button("⬇️ Download Excel", data=excel_bytes, file_name="Gefaehrdungsbeurteilung.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="btn_dl_excel")

    st.markdown("---")
    st.subheader("JSON laden")
    up = st.file_uploader("Bestehende Beurteilung (.json)", type=["json"], key="uploader_json")
    if up is not None:
        content = up.read().decode("utf-8")
        st.session_state.assessment = from_json(content)
        if not getattr(st.session_state.assessment, "industry", None):
            st.session_state.assessment.industry = "Hotel/Gastgewerbe"
        st.success("Beurteilung geladen.")
        st.rerun()

# Tabs
tabs = st.tabs([
    "0 Vorlagen auswählen", "1 Vorbereiten", "2 Ermitteln", "3 Beurteilen", "4 Maßnahmen",
    "5 Umsetzen", "6 Wirksamkeit", "7 Dokumentation", "8 Fortschreiben", "Übersicht"
])

# 0 Vorlagen auswählen
with tabs[0]:
    st.subheader("0) Vorlagen auswählen (Tätigkeiten/Gefährdungen per Häkchen übernehmen)")
    st.caption("Branche wählen, filtern, Häkchen setzen, dann übernehmen. Mehrfach-Gefährdungen werden – wenn Option aktiv – automatisch in Einzel-Gefährdungen getrennt.")

    lib = INDUSTRY_LIBRARY.get(sector, {})
    all_areas = list(lib.keys())
    area_filter = st.multiselect("Bereiche filtern", options=all_areas, default=all_areas, key="tmpl_area_filter")
    text_filter = st.text_input("Textfilter (Activity/Gefährdung enthält…)", key="tmpl_text_filter").strip().lower()

    if "template_checks" not in st.session_state:
        st.session_state.template_checks = {}

    cols = st.columns([0.24, 0.24, 0.42, 0.10])
    cols[0].markdown("**Bereich**")
    cols[1].markdown("**Tätigkeit**")
    cols[2].markdown("**Gefährdung**")
    cols[3].markdown("**Auswählen**")

    items = iter_template_items(sector)
    shown_keys = []
    for area, item, keyval in items:
        if area_filter and area not in area_filter:
            continue
        if text_filter:
            blob = f"{item.get('activity','')} {item.get('hazard','')}".lower()
            if text_filter not in blob:
                continue
        shown_keys.append(keyval)
        c0, c1, c2, c3 = st.columns([0.24, 0.24, 0.42, 0.10])
        c0.write(area)
        c1.write(item.get("activity",""))
        c2.write(item.get("hazard",""))
        default_checked = st.session_state.template_checks.get(keyval, False)
        st.session_state.template_checks[keyval] = c3.checkbox(" ", key=f"chk_{keyval}", value=default_checked)

    st.markdown("---")
    colA, colB, colC = st.columns([0.5,0.25,0.25])
    with colB:
        if st.button("Alle sichtbaren markieren", key="btn_mark_all"):
            for k in shown_keys:
                st.session_state.template_checks[k] = True
            st.rerun()
    with colC:
        if st.button("Alle sichtbaren demarkieren", key="btn_unmark_all"):
            for k in shown_keys:
                st.session_state.template_checks[k] = False
            st.rerun()

    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        if st.button("➕ Ausgewählte übernehmen (ANHÄNGEN)", key="btn_apply_append"):
            selected = [k for k, v in st.session_state.template_checks.items() if v]
            add_template_items(assess, lib, selected_keys=selected, industry_name=sector)
            st.success(f"{len(selected)} Aktivitäten übernommen (Mehrfach-Gefährdungen ggf. aufgetrennt).")
    with col2:
        if st.button("🧹 Ausgewählte übernehmen (ERSETZEN)", key="btn_apply_replace"):
            selected = [k for k, v in st.session_state.template_checks.items() if v]
            assess.hazards = []
            add_template_items(assess, lib, selected_keys=selected, industry_name=sector)
            assess.industry = sector
            st.success(f"Vorlage ersetzt. {len(selected)} Aktivitäten übernommen (Mehrfach-Gefährdungen ggf. aufgetrennt).")
            st.rerun()

    # komplette Vorlage ohne Auswahl übernehmen (ERSETZEN)
    st.markdown("---")
    if st.button("📦 Komplette Branchenvorlage übernehmen (ERSETZEN) – ohne Auswahl", key="btn_full_template_replace"):
        assess.hazards = []
        add_template_items(assess, lib, selected_keys=None, industry_name=sector)
        assess.industry = sector
        if "template_checks" in st.session_state:
            st.session_state.template_checks = {}
        st.success(f"Komplette Vorlage '{sector}' geladen (Mehrfach-Gefährdungen ggf. aufgetrennt).")
        st.rerun()

# 1 Vorbereiten
with tabs[1]:
    st.subheader("1) Vorbereiten")
    assess.industry = st.selectbox(
        "Branche der Beurteilung", options=list(INDUSTRY_LIBRARY.keys()),
        index=list(INDUSTRY_LIBRARY.keys()).index(assess.industry) if assess.industry in INDUSTRY_LIBRARY else 0,
        key="assess_industry"
    )
    assess.scope_note = st.text_area(
        "Umfang / Arbeitsbereiche / Beteiligte",
        value=assess.scope_note, height=140, key="scope_note"
    )
    st.info("Mit Tab „0 Vorlagen auswählen“ kannst du weitere Tätigkeiten/Gefährdungen anfügen.")

# 2 Ermitteln
with tabs[2]:
    st.subheader("2) Gefährdungen ermitteln")

    # Fallback-Start, wenn noch nichts geladen
    if not assess.hazards:
        st.warning("Noch keine Gefährdungen vorhanden.")
        if st.button("🚀 Branchenvorlage jetzt laden und Beurteilung starten (ERSETZEN)", key="btn_fallback_load_from_tab2"):
            assess.hazards = []
            current_sector = st.session_state.get("sel_industry", assess.industry)
            tmpl = INDUSTRY_LIBRARY.get(current_sector, {})
            add_template_items(assess, tmpl, selected_keys=None, industry_name=current_sector)
            assess.industry = current_sector
            st.success(f"Vorlage '{assess.industry}' geladen. Du kannst jetzt beurteilen.")
            st.rerun()

    colL, colR = st.columns([2,1])

    with colL:
        st.markdown("**Gefährdungen (Bearbeiten)**")
        if assess.hazards:
            df = pd.DataFrame([hazard_to_row(h) for h in assess.hazards])
            st.dataframe(df, use_container_width=True, hide_index=True, key="df_hazards")
        else:
            st.info("Nutze den Start-Button oben, Tab 0 oder die Sidebar, um eine Vorlage zu laden.")

        with st.expander("➕ Gefährdung manuell hinzufügen"):
            col1, col2 = st.columns(2)
            known_areas = sorted({h.area for h in assess.hazards} | set(INDUSTRY_LIBRARY.get(assess.industry, {}).keys()) | {"Sonstiges"})
            area = col1.selectbox("Bereich", known_areas, key="add_area")
            activity = col2.text_input("Tätigkeit", key="add_activity")
            hazard_txt = st.text_input("Gefährdung (bei mehreren: Komma/Slash/‚und‘ trennt in Einzeleinträge)", key="add_hazard")
            sources = st.text_input("Quellen/Einwirkungen (durch ; trennen)", key="add_sources")
            existing = st.text_input("Bestehende Maßnahmen (durch ; trennen)", key="add_existing")
            if st.button("Hinzufügen", key="btn_add_hazard"):
                hazards_list = split_hazard_text(hazard_txt) if st.session_state.get("opt_split_multi_hazards", True) else [hazard_txt]
                for hz_text in hazards_list:
                    assess.hazards.append(Hazard(
                        id=new_id(), area=area, activity=activity, hazard=hz_text,
                        sources=[s.strip() for s in sources.split(";") if s.strip()],
                        existing_controls=[e.strip() for e in existing.split(";") if e.strip()]
                    ))
                st.success(f"{len(hazards_list)} Eintrag(e) hinzugefügt (1 Tätigkeit → 1 Gefährdung je Eintrag).")

    with colR:
        st.markdown("**Auswahl & Details**")
        ids = [h.id for h in assess.hazards]
        sel_id = st.selectbox("Gefährdung auswählen (ID)", options=["--"] + ids, index=0, key="sel_hazard_edit")
        if sel_id != "--":
            hz = next(h for h in assess.hazards if h.id == sel_id)
            all_areas = list(INDUSTRY_LIBRARY.get(assess.industry, {}).keys()) + ["Sonstiges"]
            idx = all_areas.index(hz.area) if hz.area in all_areas else len(all_areas)-1
            hz.area = st.selectbox("Bereich", options=all_areas, index=idx, key=f"edit_area_{hz.id}")
            hz.activity = st.text_input("Tätigkeit", value=hz.activity, key=f"edit_activity_{hz.id}")
            hz.hazard = st.text_input("Gefährdung (nur eine pro Eintrag)", value=hz.hazard, key=f"edit_hazard_{hz.id}")
            src = st.text_area("Quellen/Einwirkungen", value="; ".join(hz.sources), key=f"edit_sources_{hz.id}")
            hz.sources = [s.strip() for s in src.split(";") if s.strip()]
            ex = st.text_area("Bestehende Maßnahmen", value="; ".join(hz.existing_controls), key=f"edit_existing_{hz.id}")
            hz.existing_controls = [e.strip() for e in ex.split(";") if e.strip()]
            if st.button("🗑️ Löschen", key=f"btn_delete_{hz.id}"):
                assess.hazards = [h for h in assess.hazards if h.id != sel_id]
                st.warning("Gefährdung gelöscht.")
                st.rerun()

# 3 Beurteilen
with tabs[3]:
    st.subheader("3) Gefährdungen beurteilen (5×5)")
    thresholds = assess.risk_matrix_thresholds["thresholds"]

    # Fallback-Start in Tab 3
    if not assess.hazards:
        st.warning("Keine Gefährdungen vorhanden. Lade eine Vorlage, um mit der Beurteilung zu starten.")
        if st.button("🚀 Branchenvorlage laden (ERSETZEN)", key="btn_fallback_load_from_tab3"):
            assess.hazards = []
            current_sector = st.session_state.get("sel_industry", assess.industry)
            tmpl = INDUSTRY_LIBRARY.get(current_sector, {})
            add_template_items(assess, tmpl, selected_keys=None, industry_name=current_sector)
            assess.industry = current_sector
            st.success(f"Vorlage '{assess.industry}' geladen.")
            st.rerun()
        st.stop()

    colA, colB = st.columns([1,1])

    with colA:
        sel = st.selectbox("Gefährdung auswählen", options=[f"{h.id} – {h.area}: {h.activity} → {h.hazard}" for h in assess.hazards], key="sel_hazard_assess")
        hz = assess.hazards[[f"{h.id} – {h.area}: {h.activity} → {h.hazard}" for h in assess.hazards].index(sel)]
        hz.prob = st.slider("Eintrittswahrscheinlichkeit (1 = sehr selten … 5 = häufig)", 1, 5, hz.prob, key=f"prob_{hz.id}")
        hz.sev = st.slider("Schadensschwere (1 = gering … 5 = katastrophal)", 1, 5, hz.sev, key=f"sev_{hz.id}")
        v, lvl = compute_risk(hz.prob, hz.sev, thresholds)
        hz.risk_value, hz.risk_level = v, lvl
        color = "green" if lvl == "niedrig" else "orange" if lvl == "mittel" else "red"
        st.markdown(f"**Risikosumme:** {v}  —  **Stufe:** :{color}_circle: {lvl}")
        hz.documentation_note = st.text_area("Beurteilungs-/Dokumentationshinweis", value=hz.documentation_note, key=f"doc_note_{hz.id}")

    with colB:
        st.markdown("**Schnellübersicht (Top-Risiken)**")
        top = sorted(assess.hazards, key=lambda x: x.risk_value, reverse=True)[:10]
        top_df = pd.DataFrame([{"ID":h.id, "Bereich":h.area, "Tätigkeit":h.activity, "Gefährdung":h.hazard, "Risiko":h.risk_value, "Stufe":h.risk_level} for h in top])
        st.dataframe(top_df, hide_index=True, use_container_width=True, key="df_top_risks")

# 4 Maßnahmen
with tabs[4]:
    st.subheader("4) Maßnahmen festlegen (STOP + Q)")
    st.caption("Erst S (Quelle vermeiden/ersetzen), dann T, O, P und Q.")
    if not assess.hazards:
        st.info("Keine Gefährdungen vorhanden. Lade eine Vorlage in Tab 0 oder nutze die Sidebar.")
    else:
        sel = st.selectbox("Gefährdung auswählen", options=[f"{h.id} – {h.area}: {h.activity} → {h.hazard}" for h in assess.hazards], key="sel_hazard_measures")
        hz = assess.hazards[[f"{h.id} – {h.area}: {h.activity} → {h.hazard}" for h in assess.hazards].index(sel)]

        with st.expander("➕ Maßnahme hinzufügen"):
            title = st.text_input("Maßnahme", key=f"m_title_{hz.id}")
            stop = st.selectbox("STOP(+Q)", STOP_LEVELS, index=0, key=f"m_stop_{hz.id}")
            responsible = st.text_input("Verantwortlich", key=f"m_resp_{hz.id}")
            due = st.date_input("Fällig am", value=date.today()+relativedelta(months=1), key=f"m_due_{hz.id}")
            notes = st.text_area("Hinweis", key=f"m_note_{hz.id}")
            if st.button("Hinzufügen ➕", key=f"btn_add_measure_{hz.id}"):
                hz.additional_measures.append(Measure(title=title, stop_level=stop, responsible=responsible, due_date=due.isoformat(), notes=notes))
                st.success("Maßnahme hinzugefügt.")

        if hz.additional_measures:
            mdf = pd.DataFrame([asdict(m) for m in hz.additional_measures])
            st.dataframe(mdf, use_container_width=True, hide_index=True, key=f"df_measures_{hz.id}")

# 5 Umsetzen
with tabs[5]:
    st.subheader("5) Maßnahmen umsetzen (Plan/Status)")
    rows = []
    for h in assess.hazards:
        for m in h.additional_measures:
            rows.append({"ID": h.id, "Bereich": h.area, "Tätigkeit": h.activity, "Gefährdung": h.hazard, "Risiko": h.risk_value,
                         "Maßnahme": m.title, "STOP(+Q)": m.stop_level, "Fällig": m.due_date or "",
                         "Status": m.status, "Verantwortlich": m.responsible})
    if rows:
        plan = pd.DataFrame(rows).sort_values(by=["Risiko"], ascending=False)
        st.dataframe(plan, use_container_width=True, hide_index=True, key="df_plan")
    else:
        st.info("Noch keine Maßnahmen geplant.")

# 6 Wirksamkeit
with tabs[6]:
    st.subheader("6) Wirksamkeit überprüfen")
    if not assess.hazards:
        st.info("Keine Gefährdungen vorhanden.")
    else:
        sel = st.selectbox("Gefährdung auswählen", options=[f"{h.id} – {h.area}: {h.activity} → {h.hazard}" for h in assess.hazards], key="sel_hazard_review")
        hz = assess.hazards[[f"{h.id} – {h.area}: {h.activity} → {h.hazard}" for h in assess.hazards].index(sel)]
        if hz.additional_measures:
            for i, m in enumerate(hz.additional_measures):
                st.markdown(f"**{i+1}. {m.title}**  ({m.stop_level})")
                m.status = st.selectbox("Status", STATUS_LIST, index=STATUS_LIST.index(m.status) if m.status in STATUS_LIST else 0, key=f"stat_{hz.id}_{i}")
                m.notes = st.text_area("Wirksamkeits-/Prüfhinweis", value=m.notes, key=f"notes_{hz.id}_{i}")
        else:
            st.info("Für diese Gefährdung sind noch keine Maßnahmen hinterlegt.")
        hz.last_review = st.date_input("Datum der Überprüfung", value=date.today(), key=f"rev_date_{hz.id}").isoformat()
        hz.reviewer = st.text_input("Prüfer/in", value=hz.reviewer, key=f"rev_reviewer_{hz.id}")

# 7 Dokumentation
with tabs[7]:
    st.subheader("7) Ergebnisse dokumentieren")
    assess.documentation_note = st.text_area(
        "Dokumentationshinweis (welche Unterlagen, wo abgelegt, Versionierung)",
        value=assess.documentation_note, height=120, key="doc_note_global"
    )
    st.markdown("**Nachweise/Beispiele:** Betriebsanweisungen, Unterweisungsnachweise, Prüfprotokolle (Leitern/Elektro), Wartungspläne (z. B. Lüftung/Legionellen), Gefahrstoffverzeichnis, Unfallstatistik, Beinahe-Unfälle.")

# 8 Fortschreiben
with tabs[8]:
    st.subheader("8) Fortschreiben")
    assess.next_review_hint = st.text_area(
        "Anlässe/Fristen (regelmäßige Überprüfung, nach Unfällen/Beinaheunfällen, Änderungen)",
        value=assess.next_review_hint, height=100, key="next_review_hint"
    )
    st.info("Hinweis: Änderungen dokumentieren und Datums-/Namensfeld bei Überprüfung ergänzen.")

# Übersicht
with tabs[9]:
    st.subheader("Übersicht & Kennzahlen")
    total = len(assess.hazards)
    high = len([h for h in assess.hazards if h.risk_level in ("hoch", "sehr hoch")])
    st.metric("Gefährdungen gesamt", total)
    st.metric("Davon hoch/sehr hoch", high)
    if total:
        by_area = pd.DataFrame(pd.Series([h.area for h in assess.hazards]).value_counts(), columns=["Anzahl"])
        st.markdown("**Gefährdungen je Bereich**")
        st.dataframe(by_area, use_container_width=True, key="df_by_area")
    st.markdown("**Hinweise**")
    assess.measures_plan_note = st.text_area("Projekt-/Maßnahmenplan (kurz)", value=assess.measures_plan_note, key="measures_plan_note")
